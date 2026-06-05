"""PDPC enforcement-decision ingestion (SGLB-01 PDPA-Outcome).

Reads ``backend/data/raw/pdpc_decisions.xlsx`` (vendored from
`kevanwee/pdpcscraper`), mechanically extracts obligation tags + penalty
band from PDPC's own published rows, redacts outcome-leaking text from
the case description to produce ``fact_summary``, partitions by
publication date into ``train`` / ``dev`` / ``test`` and emits:

  - ``backend/data/benchmarks/sglb_01_pdpa/{train,dev,test}.jsonl``
    (one row per case; stable IDs; source URL + version metadata)
  - ``backend/benchmark/datasets/sglb_01_pdpa.yaml``
    (harness-compatible concatenation; the eval CLI reads this)

Mechanical-extraction policy (coverage-matrix §4.1):

  - Obligations come from PDPC's published ``Obligations`` column verbatim,
    canonicalised to a closed taxonomy.
  - Penalty band is log-bucketed from the published SGD figure using
    documented boundaries; no judgement.
  - Fact summary is the PDPC-published case description with mechanical
    redaction of outcome leakage (penalty amounts + outcome verbs).

CLI:

    python -m data.ingestion.pdpc --xlsx backend/data/raw/pdpc_decisions.xlsx \\
                                   --output backend/data/benchmarks/sglb_01_pdpa \\
                                   --yaml backend/benchmark/datasets/sglb_01_pdpa.yaml
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Iterator

import yaml

from data.ingestion._provenance import extraction_rule_sha

DATASET_VERSION = "sglb-01-v0.1"  # bump when ingestion logic changes
EXTRACTION_MODULE = Path(__file__)
EXTRACTION_RULE_NAME = "pdpc"

# Closed PDPC obligation taxonomy (incl. Accountability — observed in xlsx
# but missing from the original SGLB-01 spec draft).
OBLIGATION_TAXONOMY: tuple[str, ...] = (
    "consent",
    "notification",
    "purpose_limitation",
    "protection",
    "retention_limitation",
    "data_portability",
    "dpo",
    "dnc",
    "data_intermediary",
    "transfer_limitation",
    "accountability",
    "openness",
    "accuracy",
    "access_correction",
)

# Map xlsx free-text labels → taxonomy slugs.
_OBLIGATION_ALIASES: dict[str, str] = {
    "consent": "consent",
    "notification": "notification",
    "purpose limitation": "purpose_limitation",
    "protection": "protection",
    "retention limitation": "retention_limitation",
    "retention": "retention_limitation",
    "data portability": "data_portability",
    "dpo": "dpo",
    "data protection officer": "dpo",
    "dnc": "dnc",
    "do not call": "dnc",
    "data intermediary": "data_intermediary",
    "transfer limitation": "transfer_limitation",
    "transfer": "transfer_limitation",
    "accountability": "accountability",
    "openness": "openness",
    "accuracy": "accuracy",
    "access and correction": "access_correction",
    "access": "access_correction",
    "correction": "access_correction",
}

# Penalty band boundaries (SGD, log10-spaced).
# Reported in spec + dataset metadata for reproducibility.
PENALTY_BANDS: tuple[str, ...] = ("none", "low", "mid", "high")
_BAND_LOW_MAX = 5_000
_BAND_MID_MAX = 50_000

# Outcome-leakage redaction patterns. Applied to the description in order.
_REDACTORS: tuple[tuple[re.Pattern[str], str], ...] = (
    # Strip dollar amounts everywhere
    (re.compile(r"(?:S\$|\$)\s*\d[\d,]*(?:\.\d{2})?", re.IGNORECASE), "[AMOUNT_REDACTED]"),
    # Replace the canonical lead sentence
    (
        re.compile(
            r"^A financial penalty of \[AMOUNT_REDACTED\] was imposed and directions were issued to ",
            re.IGNORECASE,
        ),
        "An enforcement outcome resulted for ",
    ),
    (
        re.compile(
            r"^A financial penalty of \[AMOUNT_REDACTED\] was imposed on ",
            re.IGNORECASE,
        ),
        "An enforcement outcome resulted for ",
    ),
    (
        re.compile(r"\bdirections were issued\b", re.IGNORECASE),
        "regulatory action was taken",
    ),
    (
        re.compile(r"\bfinancial penalt(?:y|ies)\b", re.IGNORECASE),
        "enforcement outcome",
    ),
    (
        re.compile(r"\bClick here for more information\.?", re.IGNORECASE),
        "",
    ),
    # Collapse runs of whitespace introduced by redaction
    (re.compile(r"\s{2,}"), " "),
)

# Date split boundaries.
# Coverage matrix §4.3 sets the contamination cutoff at 2026-Q1.
_TEST_START = dt.date(2026, 1, 1)
_DEV_START = dt.date(2024, 1, 1)

_DATE_FORMATS = ("%d %b %Y", "%d %B %Y", "%Y-%m-%d")


@dataclass
class PdpcCase:
    case_id: str
    case_name: str
    citation: str
    pub_date: dt.date
    obligations: list[str]
    penalty_band: str
    penalty_sgd: int | None
    fact_summary: str
    source_url: str
    decision_type: str
    split: str = ""
    raw_obligations: str = ""
    raw_penalty: str = ""
    excluded_reason: str = ""

    def as_jsonl_row(self, rule_sha: str | None = None) -> dict:
        sha = rule_sha or extraction_rule_sha(EXTRACTION_MODULE)
        return {
            "id": self.case_id,
            "extraction_rule_sha": sha,
            "inputs": {"fact_summary": self.fact_summary},
            "expected_output": {
                "obligations": self.obligations,
                "penalty_band": self.penalty_band,
            },
            "metadata": {
                "task": "SGLB-01",
                "split": self.split,
                "jurisdiction": "SG",
                "case_name": self.case_name,
                "citation": self.citation,
                "pub_date": self.pub_date.isoformat(),
                "penalty_sgd": self.penalty_sgd,
                "decision_type": self.decision_type,
                "source_url": self.source_url,
                "source_adapter": "api.adapters.public.pdpc.PdpcAdapter",
                "dataset_version": DATASET_VERSION,
                "raw_obligations": self.raw_obligations,
                "raw_penalty": self.raw_penalty,
                "label_provenance": "mechanical-extraction-from-pdpc-published-row",
            },
        }


@dataclass
class IngestStats:
    total: int = 0
    written: int = 0
    excluded: list[tuple[str, str]] = field(default_factory=list)
    by_split: dict[str, int] = field(default_factory=dict)
    band_distribution: dict[str, int] = field(default_factory=dict)


def stable_id(source_url: str) -> str:
    digest = hashlib.sha256(source_url.encode("utf-8")).hexdigest()[:12]
    return f"sglb_01_{digest}"


def parse_pub_date(raw: str) -> dt.date | None:
    s = (raw or "").strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def canonical_obligations(raw: str) -> list[str]:
    if not raw:
        return []
    seen: list[str] = []
    for part in str(raw).split(","):
        key = part.strip().lower()
        slug = _OBLIGATION_ALIASES.get(key)
        if slug is None or slug not in OBLIGATION_TAXONOMY:
            continue
        if slug not in seen:
            seen.append(slug)
    return seen


def parse_penalty_sgd(raw: str) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() == "none":
        return None
    match = re.search(r"([\d,]+(?:\.\d{2})?)", s)
    if not match:
        return None
    try:
        return int(float(match.group(1).replace(",", "")))
    except ValueError:
        return None


def derive_penalty_band(amount: int | None) -> str:
    if amount is None or amount <= 0:
        return "none"
    if amount < _BAND_LOW_MAX:
        return "low"
    if amount < _BAND_MID_MAX:
        return "mid"
    return "high"


def redact_fact_summary(description: str) -> str:
    text = (description or "").strip()
    for pattern, repl in _REDACTORS:
        text = pattern.sub(repl, text)
    return text.strip()


def assign_split(pub_date: dt.date) -> str:
    if pub_date >= _TEST_START:
        return "test"
    if pub_date >= _DEV_START:
        return "dev"
    return "train"


def _iter_xlsx_rows(xlsx_path: Path) -> Iterator[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to ingest PDPC; add it to backend/pyproject.toml"
        ) from exc
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    keys = [str(c or "").strip() for c in header]
    for row in rows:
        if row is None:
            continue
        yield dict(zip(keys, row))


def _row_to_case(row: dict) -> tuple[PdpcCase | None, str]:
    """Return (case, excluded_reason). On exclusion, case is None."""
    url = str(row.get("URL") or "").strip()
    if not url:
        return None, "missing url"

    description = str(row.get("Case Description") or "")
    fact_summary = redact_fact_summary(description)
    if len(fact_summary) < 50:
        return None, "fact summary too short post-redaction"

    pub_date = parse_pub_date(str(row.get("Date") or ""))
    if pub_date is None:
        return None, "missing or unparseable date"

    raw_obs = str(row.get("Obligations") or "")
    obligations = canonical_obligations(raw_obs)
    if not obligations:
        return None, f"no obligations in taxonomy: {raw_obs!r}"

    raw_pen = str(row.get("Financial Penalty") or "")
    amount = parse_penalty_sgd(raw_pen)
    band = derive_penalty_band(amount)

    case = PdpcCase(
        case_id=stable_id(url),
        case_name=str(row.get("Case Name") or "").strip(),
        citation=str(row.get("Case Citation") or "").strip(),
        pub_date=pub_date,
        obligations=obligations,
        penalty_band=band,
        penalty_sgd=amount,
        fact_summary=fact_summary,
        source_url=url,
        decision_type=str(row.get("Decision Type") or "").strip(),
        raw_obligations=raw_obs.strip(),
        raw_penalty=raw_pen.strip(),
    )
    case.split = assign_split(pub_date)
    return case, ""


def load_cases(xlsx_path: Path) -> tuple[list[PdpcCase], list[tuple[str, str]]]:
    cases: list[PdpcCase] = []
    excluded: list[tuple[str, str]] = []
    seen_ids: set[str] = set()
    for row in _iter_xlsx_rows(xlsx_path):
        case, reason = _row_to_case(row)
        if case is None:
            excluded.append((str(row.get("URL") or row.get("Case Name") or ""), reason))
            continue
        if case.case_id in seen_ids:
            excluded.append((case.case_id, "duplicate id"))
            continue
        seen_ids.add(case.case_id)
        cases.append(case)
    # deterministic ordering by pub_date then id
    cases.sort(key=lambda c: (c.pub_date, c.case_id))
    return cases, excluded


def write_jsonl(cases: Iterable[PdpcCase], output_dir: Path) -> dict[str, int]:
    output_dir.mkdir(parents=True, exist_ok=True)
    by_split: dict[str, list[PdpcCase]] = {"train": [], "dev": [], "test": []}
    rule_sha = extraction_rule_sha(EXTRACTION_MODULE)
    for case in cases:
        by_split.setdefault(case.split, []).append(case)
    counts: dict[str, int] = {}
    for split, items in by_split.items():
        path = output_dir / f"{split}.jsonl"
        with path.open("w", encoding="utf-8") as fp:
            for case in items:
                fp.write(json.dumps(case.as_jsonl_row(rule_sha), sort_keys=True) + "\n")
        counts[split] = len(items)
    return counts


def write_harness_yaml(cases: Iterable[PdpcCase], yaml_path: Path) -> int:
    yaml_path.parent.mkdir(parents=True, exist_ok=True)
    case_list = list(cases)
    rule_sha = extraction_rule_sha(EXTRACTION_MODULE)
    payload = {
        "extraction_rules": {EXTRACTION_RULE_NAME: rule_sha},
        "cases": [
            {
                "name": c.case_id,
                "extraction_rule_sha": rule_sha,
                "inputs": {"fact_summary": c.fact_summary},
                "expected_output": {
                    "obligations": c.obligations,
                    "penalty_band": c.penalty_band,
                },
                "metadata": {
                    "task": "SGLB-01",
                    "split": c.split,
                    "jurisdiction": "SG",
                    "pub_date": c.pub_date.isoformat(),
                    "citation": c.citation,
                    "source_url": c.source_url,
                    "dataset_version": DATASET_VERSION,
                },
            }
            for c in case_list
        ]
    }
    yaml_path.write_text(
        yaml.safe_dump(payload, sort_keys=False, default_flow_style=False, width=120, allow_unicode=True),
        encoding="utf-8",
    )
    return len(case_list)


def ingest(*, xlsx_path: Path, output_dir: Path, yaml_path: Path) -> IngestStats:
    cases, excluded = load_cases(xlsx_path)
    by_split = write_jsonl(cases, output_dir)
    write_harness_yaml(cases, yaml_path)
    bands: dict[str, int] = {}
    for c in cases:
        bands[c.penalty_band] = bands.get(c.penalty_band, 0) + 1
    return IngestStats(
        total=len(cases) + len(excluded),
        written=len(cases),
        excluded=excluded,
        by_split=by_split,
        band_distribution=bands,
    )


def _default_repo_root() -> Path:
    # __file__ = backend/data/ingestion/pdpc.py → parents[3] = repo root
    return Path(__file__).resolve().parents[3]


def main(argv: list[str] | None = None) -> int:
    root = _default_repo_root()
    parser = argparse.ArgumentParser(prog="data.ingestion.pdpc", description="Ingest PDPC enforcement decisions for SGLB-01")
    parser.add_argument(
        "--xlsx",
        default=str(root / "backend" / "data" / "raw" / "pdpc_decisions.xlsx"),
        help="path to pdpc_decisions.xlsx (default: backend/data/raw/)",
    )
    parser.add_argument(
        "--output",
        default=str(root / "backend" / "data" / "benchmarks" / "sglb_01_pdpa"),
        help="output dir for {train,dev,test}.jsonl",
    )
    parser.add_argument(
        "--yaml",
        default=str(root / "backend" / "benchmark" / "datasets" / "sglb_01_pdpa.yaml"),
        help="harness YAML dataset path",
    )
    args = parser.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"error: xlsx not found: {xlsx_path}", file=sys.stderr)
        return 2

    stats = ingest(xlsx_path=xlsx_path, output_dir=Path(args.output), yaml_path=Path(args.yaml))
    payload = {
        "total_seen": stats.total,
        "written": stats.written,
        "excluded": len(stats.excluded),
        "by_split": stats.by_split,
        "band_distribution": stats.band_distribution,
        "dataset_version": DATASET_VERSION,
    }
    print(json.dumps(payload, indent=2, sort_keys=True))
    if stats.excluded:
        print(
            f"note: excluded {len(stats.excluded)} rows; first 5 reasons: "
            + json.dumps(stats.excluded[:5]),
            file=sys.stderr,
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
